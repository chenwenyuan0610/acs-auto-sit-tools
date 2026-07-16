from io import BytesIO
import json

import pytest
from openpyxl import Workbook

from acs_auto_sit import sit_runner as sit_runner_module
from acs_auto_sit.sit_runner import load_browser_case_catalog
from acs_auto_sit.wording_profiles import (
    DEFAULT_SUPPORTED_LOCALES,
    build_localized_wording_cases,
    import_wording_workbook,
    load_wording_profiles,
)


ISSUER_HEADERS = (
    "啟用",
    "發卡行代號",
    "發卡行名稱",
    "Issuer Mode",
    "預設語言",
    "OTP 來源",
    "成功 OTP",
    "失敗 OTP",
    "備註",
)
WORDING_HEADERS = (
    "啟用匯入",
    "發卡行代號",
    "發卡行名稱",
    "Issuer Mode",
    "設備通道",
    "訊息類別",
    "來源頁籤",
    "編碼代號",
    "語言代碼",
    "欄位代號",
    "欄位名稱",
    "選項序號",
    "話術內容",
    "Placeholder",
    "完成狀態",
    "備註",
)
RAW_WORDING_HEADERS = (
    "設備通道",
    "訊息類別",
    "編碼代號",
    "語言代碼",
    "驗證訊息標題",
    "驗證訊息欄位文字",
    "驗證訊息欄位標籤",
    "第二組驗證碼標籤",
    "下一步標籤",
    "重送驗證標籤",
    "繼續OOB作業標籤",
    "是否需要幫助標籤",
    "幫助文字",
    "是否完成",
)


def _workbook_bytes(wording_rows=(), *, include_wording_sheet=True):
    workbook = Workbook()
    issuer_sheet = workbook.active
    issuer_sheet.title = "發卡行設定"
    issuer_sheet.append(ISSUER_HEADERS)
    issuer_sheet.append(
        ("Y", "default", "預設發卡行", "direct_otp", "zh_TW", "客戶產客戶驗", "123456", "000000", "")
    )
    issuer_sheet.append(("Y", None, None, "selection_sms_otp", "zh_TW", "我們產我們驗", None, None, ""))

    if include_wording_sheet:
        wording_sheet = workbook.create_sheet("話術匯入")
        wording_sheet.append(WORDING_HEADERS)
        for row in wording_rows:
            wording_sheet.append(row)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _wording_row(locale, field_key, content, *, issuer_id=None):
    return (
        "Y",
        issuer_id,
        None,
        "direct_otp",
        "BROWSER",
        "PA",
        "SMS",
        "SEND_SMS_OTP",
        locale,
        field_key,
        field_key,
        None,
        content,
        None,
        "Pass",
        None,
    )


def _raw_workbook_bytes(
    *,
    duplicate_sms=False,
    conflicting_sms=False,
    omit_email_column=None,
    metadata_only_email_row=False,
):
    workbook = Workbook()
    workbook.remove(workbook.active)
    source_rows = {
        "SMS": (
            "BROWSER", "PA", "SEND_SMS_OTP", "zh_TW", "簡訊驗證", "請輸入驗證碼 {0}<br>完成交易",
            "驗證碼", "確認驗證碼", "下一步", "重送驗證碼", "", "需要幫助？", "請聯絡客服", "Y",
        ),
        "Email": (
            "BROWSER", "PA", "SEND_EMAIL_OTP", "en_US", "Email verification", "Enter the code sent by email",
            "Verification code", "Confirm code", "Next", "Resend email", "", "Need help?", "Contact support", "Y",
        ),
        "OOB": (
            "BROWSER", "NPA", "OOB_AUTHENTICATION", "zh_CN", "交易验证", "请在应用程序中确认交易",
            "", "", "", "", "继续验证", "需要帮助？", "请联系客户服务", "N",
        ),
        "Single Select": (
            "BROWSER", "PA", "SELECT_AUTHENTICATION_METHOD", "zh_TW", "選擇驗證方式", "請選擇驗證方式",
            "驗證方式", "", "繼續", "", "", "需要幫助？", "請聯絡客服", "Y",
        ),
    }
    for sheet_name, row in source_rows.items():
        headers = list(RAW_WORDING_HEADERS)
        if sheet_name == "Email" and omit_email_column:
            headers.remove(omit_email_column)
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        sheet.append(tuple(value for index, value in enumerate(row) if RAW_WORDING_HEADERS[index] in headers))
        if sheet_name == "SMS" and duplicate_sms:
            sheet.append(row)
        if sheet_name == "SMS" and conflicting_sms:
            changed = list(row)
            changed[4] = "不同的簡訊驗證標題"
            sheet.append(changed)
        if sheet_name == "Email" and metadata_only_email_row:
            metadata_row = [None] * len(headers)
            metadata_row[headers.index("是否完成")] = "pass"
            sheet.append(metadata_row)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_import_normalizes_and_persists_workbook(tmp_path):
    destination = tmp_path / "wording_profiles.json"
    content = _workbook_bytes(
        (
            _wording_row("zh_TW", "challenge_title", "交易驗證"),
            _wording_row("en_US", "challenge_title", "Transaction Verification"),
            _wording_row("zh_CN", "challenge_title", "交易验证"),
        )
    )

    imported = import_wording_workbook(content, destination, source_file="issuer-wording.xlsx")

    assert destination.is_file()
    assert imported["defaultSupportedLocales"] == list(DEFAULT_SUPPORTED_LOCALES)
    assert imported["issuers"]["default"]["issuerModes"] == ["direct_otp", "selection_sms_otp"]
    assert imported["issuers"]["default"]["supportedLocales"] == list(DEFAULT_SUPPORTED_LOCALES)
    assert imported["wordings"][0] == {
        "issuerId": "default",
        "issuerMode": "direct_otp",
        "deviceChannel": "BROWSER",
        "messageCategory": "PA",
        "sourceSheet": "SMS",
        "wordingCode": "SEND_SMS_OTP",
        "locale": "zh_TW",
        "fieldKey": "challenge_title",
        "content": "交易驗證",
        "placeholders": [],
        "completionStatus": "Pass",
    }
    assert imported["sourceFormat"] == "normalized"
    assert imported["sourceSheets"] == ["SMS"]
    assert imported["summary"] == {
        "issuerCount": 1,
        "localeCount": 3,
        "wordingCount": 3,
        "sourceSheetCount": 1,
        "normalizedRowCount": 3,
    }
    assert load_wording_profiles(destination) == json.loads(destination.read_text(encoding="utf-8"))


def test_import_rejects_missing_required_sheet(tmp_path):
    with pytest.raises(ValueError, match="話術匯入"):
        import_wording_workbook(
            _workbook_bytes(include_wording_sheet=False),
            tmp_path / "wording_profiles.json",
        )


def test_import_rejects_duplicate_wording_key_without_replacing_valid_file(tmp_path):
    destination = tmp_path / "wording_profiles.json"
    destination.write_text('{"version": 0}', encoding="utf-8")
    duplicate = _wording_row("en_US", "challenge_title", "Transaction Verification")
    conflict = _wording_row("en_US", "challenge_title", "Different title")

    with pytest.raises(ValueError, match="Duplicate wording key"):
        import_wording_workbook(
            _workbook_bytes((duplicate, conflict)),
            destination,
        )

    assert json.loads(destination.read_text(encoding="utf-8")) == {"version": 0}


def test_import_detects_and_normalizes_raw_challenge_ui_workbook(tmp_path):
    destination = tmp_path / "wording_profiles.json"

    imported = import_wording_workbook(
        _raw_workbook_bytes(duplicate_sms=True),
        destination,
        source_file="challenge_ui_info.xlsx",
    )

    assert imported["sourceFormat"] == "challenge_ui_info"
    assert imported["sourceSheets"] == ["SMS", "Email", "OOB", "Single Select"]
    assert imported["defaultSupportedLocales"] == list(DEFAULT_SUPPORTED_LOCALES)
    assert imported["issuers"]["default"]["supportedLocales"] == list(DEFAULT_SUPPORTED_LOCALES)
    assert imported["summary"]["sourceSheetCount"] == 4
    assert imported["summary"]["normalizedRowCount"] == 4
    assert imported["summary"]["wordingCount"] == 27

    wordings = imported["wordings"]
    sms_message = next(
        item
        for item in wordings
        if item["sourceSheet"] == "SMS"
        and item["wordingCode"] == "SEND_SMS_OTP"
        and item["fieldKey"] == "challenge_message"
    )
    assert sms_message["content"] == "請輸入驗證碼 {0}<br>完成交易"
    assert sms_message["placeholders"] == ["{0}"]
    assert sms_message["completionStatus"] == "Y"

    oob_continue = next(
        item for item in wordings if item["sourceSheet"] == "OOB" and item["fieldKey"] == "continue_oob_button"
    )
    assert oob_continue["content"] == "继续验证"
    assert load_wording_profiles(destination) == imported


def test_import_raw_workbook_rejects_missing_required_source_column_without_replacing_valid_file(tmp_path):
    destination = tmp_path / "wording_profiles.json"
    destination.write_text('{"version": 0}', encoding="utf-8")

    with pytest.raises(ValueError, match=r"Sheet Email is missing column\(s\): 語言代碼"):
        import_wording_workbook(
            _raw_workbook_bytes(omit_email_column="語言代碼"),
            destination,
        )

    assert json.loads(destination.read_text(encoding="utf-8")) == {"version": 0}


def test_import_raw_workbook_rejects_conflicting_duplicate_runtime_key(tmp_path):
    with pytest.raises(ValueError, match="Duplicate wording key"):
        import_wording_workbook(
            _raw_workbook_bytes(conflicting_sms=True),
            tmp_path / "wording_profiles.json",
        )


def test_import_raw_workbook_ignores_metadata_only_row(tmp_path):
    imported = import_wording_workbook(
        _raw_workbook_bytes(metadata_only_email_row=True),
        tmp_path / "wording_profiles.json",
    )

    assert imported["summary"]["normalizedRowCount"] == 4
    assert imported["summary"]["wordingCount"] == 27


def _normalized_profiles(*, omit_code=None):
    scenarios = (
        ("PA", "SEND_SMS_OTP"),
        ("NPA", "SEND_SMS_OTP"),
        ("PA", "INCORRECT_SMS_OTP"),
        ("PA", "RESEND_SMS_OTP"),
        ("PA", "RESEND_SMS_GAP_LIMIT"),
        ("PA", "RESEND_SMS_LIMIT_EXCEED"),
        ("PA", "SMS_PASSCODE_EXPIRED"),
    )
    wordings = []
    for category, code in scenarios:
        if code == omit_code:
            continue
        for locale in DEFAULT_SUPPORTED_LOCALES:
            for field_key, content in (
                ("challenge_title", f"{code} {locale}"),
                ("challenge_message", f"{code} message {locale}"),
            ):
                wordings.append(
                    {
                        "issuerId": "default",
                        "issuerMode": "direct_otp",
                        "deviceChannel": "BROWSER",
                        "messageCategory": category,
                        "wordingCode": code,
                        "locale": locale,
                        "fieldKey": field_key,
                        "content": content,
                        "placeholders": [],
                    }
                )
    return {
        "defaultSupportedLocales": list(DEFAULT_SUPPORTED_LOCALES),
        "issuers": {
            "default": {
                "id": "default",
                "name": "Default Issuer",
                "defaultLocale": "zh_TW",
                "issuerModes": ["direct_otp"],
                "supportedLocales": list(DEFAULT_SUPPORTED_LOCALES),
            }
        },
        "wordings": wordings,
    }


def _source_templates():
    return [
        {
            "id": case_id,
            "functionPoint": f"Legacy {case_id}",
            "steps": ["Open challenge"],
            "expected": {"prompts": ["legacy"]},
            "automation": {"status": "automatable", "tags": ["otp", "pa"]},
        }
        for case_id in ("case23", "case27", "case31", "case35", "case39", "case43", "case47")
    ]


def test_build_localized_cases_expands_seven_scenarios_for_default_locales():
    cases = build_localized_wording_cases(_normalized_profiles(), _source_templates())

    assert len(cases) == 21
    assert [case["id"] for case in cases[:3]] == ["case23_zh_TW", "case23_en_US", "case23_zh_CN"]
    assert cases[0]["baseCaseId"] == "case23"
    assert cases[0]["locale"] == "zh_TW"
    assert cases[0]["browserLanguage"] == "zh-TW"
    assert cases[0]["wording"]["code"] == "SEND_SMS_OTP"
    assert cases[0]["expected"]["uiFields"] == {
        "challenge_title": "SEND_SMS_OTP zh_TW",
        "challenge_message": "SEND_SMS_OTP message zh_TW",
    }
    assert cases[0]["expected"]["prompts"] == [
        "SEND_SMS_OTP zh_TW",
        "SEND_SMS_OTP message zh_TW",
    ]
    assert cases[0]["expected"]["validationMode"] == "excel_fields"
    assert cases[0]["availability"] == {"enabled": True, "reason": ""}


def test_build_localized_cases_disables_missing_wording_scenario():
    cases = build_localized_wording_cases(
        _normalized_profiles(omit_code="RESEND_SMS_LIMIT_EXCEED"),
        _source_templates(),
    )

    missing = [case for case in cases if case["baseCaseId"] == "case43"]
    assert len(missing) == 3
    assert all(case["availability"]["enabled"] is False for case in missing)
    assert all("RESEND_SMS_LIMIT_EXCEED" in case["availability"]["reason"] for case in missing)


def test_build_localized_cases_prefers_issuer_wording_over_shared_default():
    profiles = _normalized_profiles()
    profiles["issuers"]["bank_a"] = {
        "id": "bank_a",
        "name": "Bank A",
        "defaultLocale": "zh_TW",
        "issuerModes": ["direct_otp"],
        "supportedLocales": list(DEFAULT_SUPPORTED_LOCALES),
    }
    profiles["wordings"].append(
        {
            "issuerId": "bank_a",
            "issuerMode": "direct_otp",
            "deviceChannel": "BROWSER",
            "messageCategory": "PA",
            "wordingCode": "SEND_SMS_OTP",
            "locale": "zh_TW",
            "fieldKey": "challenge_title",
            "content": "Bank A 交易驗證",
            "placeholders": [],
        }
    )

    cases = build_localized_wording_cases(profiles, _source_templates(), issuer_id="bank_a")
    cases_by_id = {case["id"]: case for case in cases}

    assert cases_by_id["case23_zh_TW"]["expected"]["prompts"][0] == "Bank A 交易驗證"
    assert cases_by_id["case23_en_US"]["expected"]["prompts"][0] == "SEND_SMS_OTP en_US"


def test_build_localized_cases_disables_incomplete_required_fields():
    profiles = _normalized_profiles()
    profiles["wordings"] = [
        wording
        for wording in profiles["wordings"]
        if not (
            wording["messageCategory"] == "PA"
            and wording["wordingCode"] == "SEND_SMS_OTP"
            and wording["locale"] == "zh_TW"
            and wording["fieldKey"] == "challenge_message"
        )
    ]

    cases = build_localized_wording_cases(profiles, _source_templates())
    case = next(case for case in cases if case["id"] == "case23_zh_TW")

    assert case["availability"]["enabled"] is False
    assert "challenge_message" in case["availability"]["reason"]


@pytest.mark.parametrize(
    ("issuer_mode", "source_sheet", "destination"),
    (
        ("sms_otp", "SMS", "sms"),
        ("email_otp", "Email", "email"),
        ("direct_oob", "OOB", "oob"),
    ),
)
def test_build_raw_localized_cases_filters_direct_mode_source_sheet(
    tmp_path,
    issuer_mode,
    source_sheet,
    destination,
):
    profiles = import_wording_workbook(
        _raw_workbook_bytes(),
        tmp_path / f"{issuer_mode}.json",
    )

    cases = build_localized_wording_cases(
        profiles,
        _source_templates(),
        issuer_mode=issuer_mode,
    )

    assert cases
    assert {case["wording"]["sourceSheet"] for case in cases} == {source_sheet}
    assert {case["flow"]["destination"] for case in cases} == {destination}
    assert all(case["flow"]["kind"] == "direct" for case in cases)
    assert all(case["expected"]["validationMode"] == "excel_fields" for case in cases)
    assert len({case["id"] for case in cases}) == len(cases)


def test_build_raw_localized_cases_generates_all_single_select_destinations(tmp_path):
    profiles = import_wording_workbook(
        _raw_workbook_bytes(),
        tmp_path / "selection.json",
    )

    cases = build_localized_wording_cases(
        profiles,
        _source_templates(),
        issuer_mode="selection_sms_email_oob",
    )

    selection_pages = [case for case in cases if case["flow"]["kind"] == "selection_page"]
    branches = [case for case in cases if case["flow"]["kind"] == "selection_branch"]
    assert len(selection_pages) == 1
    assert selection_pages[0]["flow"]["destinations"] == ["sms", "email", "oob"]
    assert {case["flow"]["destination"] for case in branches} == {"sms", "email", "oob"}
    sms_branch = next(case for case in branches if case["flow"]["destination"] == "sms")
    assert sms_branch["id"] == "ui_select_sms_pa_send_sms_otp_zh_TW"
    assert [stage["type"] for stage in sms_branch["flow"]["stages"]] == ["single_select", "sms"]
    assert sms_branch["availability"] == {"enabled": True, "reason": ""}
    email_branch = next(case for case in branches if case["flow"]["destination"] == "email")
    assert email_branch["availability"]["enabled"] is False
    assert "Single Select" in email_branch["availability"]["reason"]


def test_build_raw_localized_cases_generates_oob_to_sms_switch_metadata(tmp_path):
    profiles = import_wording_workbook(
        _raw_workbook_bytes(),
        tmp_path / "switch.json",
    )

    cases = build_localized_wording_cases(
        profiles,
        _source_templates(),
        issuer_mode="default_oob_can_switch_otp",
    )

    assert len(cases) == 1
    case = cases[0]
    assert case["flow"]["kind"] == "oob_switch_sms"
    assert case["flow"]["switchCreq"] is True
    assert [stage["type"] for stage in case["flow"]["stages"]] == ["oob", "sms"]
    assert case["availability"]["enabled"] is False
    assert "SMS" in case["availability"]["reason"]


def test_browser_catalog_replaces_legacy_localized_cases_when_profile_exists(tmp_path):
    profile_path = tmp_path / "wording_profiles.json"
    profile_path.write_text(json.dumps(_normalized_profiles()), encoding="utf-8")

    catalog = load_browser_case_catalog(
        progress_path=tmp_path / "missing-progress.json",
        wording_profiles_path=profile_path,
        issuer_id="default",
        issuer_mode="direct_otp",
    )
    case_ids = [case["id"] for case in catalog["cases"]]
    assert "case20" in case_ids
    assert "case23" not in case_ids
    assert "case23_zh_TW" in case_ids
    assert "case47_zh_CN" in case_ids
    assert len([case_id for case_id in case_ids if "_zh_" in case_id or case_id.endswith("_en_US")]) == 21
    assert catalog["wordingProfile"] == {
        "enabled": True,
        "issuerId": "default",
        "issuerMode": "direct_otp",
        "supportedLocales": list(DEFAULT_SUPPORTED_LOCALES),
    }


def test_browser_catalog_preserves_raw_case_flow_metadata(tmp_path):
    profile_path = tmp_path / "raw-wording-profiles.json"
    import_wording_workbook(_raw_workbook_bytes(), profile_path)

    catalog = load_browser_case_catalog(
        progress_path=tmp_path / "missing-progress.json",
        wording_profiles_path=profile_path,
        issuer_mode="selection_sms_email_oob",
    )

    cases = {case["id"]: case for case in catalog["cases"]}
    generated = cases["ui_select_sms_pa_send_sms_otp_zh_TW"]
    assert generated["flow"]["kind"] == "selection_branch"
    assert generated["flow"]["destination"] == "sms"
    assert generated["availability"]["enabled"] is True
    assert generated["caseImplementation"]["status"] == "completed"
    assert "case23" not in cases


def test_browser_catalog_derives_generated_capability_status(monkeypatch, tmp_path):
    profiles = _normalized_profiles()
    generated_cases = [
        {
            "id": "ui_supported",
            "wordingScenario": "incorrect_otp",
            "wording": {"code": "INCORRECT_SMS_OTP"},
            "flow": {"kind": "direct", "destination": "sms", "stages": []},
            "availability": {"enabled": True, "reason": ""},
        },
        {
            "id": "ui_pending",
            "wordingScenario": "unknown",
            "wording": {"code": "UNKNOWN"},
            "flow": {"kind": "direct", "destination": "sms", "stages": []},
            "availability": {"enabled": True, "reason": ""},
        },
    ]
    monkeypatch.setattr(sit_runner_module, "load_wording_profiles", lambda path: profiles)
    monkeypatch.setattr(
        sit_runner_module,
        "build_localized_wording_cases",
        lambda *args, **kwargs: generated_cases,
    )

    catalog = load_browser_case_catalog(
        progress_path=tmp_path / "missing-progress.json",
        wording_profiles_path=tmp_path / "wording_profiles.json",
        issuer_mode="selection_sms_email_oob",
    )
    implementations = {
        case["id"]: case["caseImplementation"]
        for case in catalog["cases"]
        if case["id"] in {"ui_supported", "ui_pending"}
    }

    assert implementations["ui_supported"]["status"] == "completed"
    assert implementations["ui_pending"]["status"] == "pending"


def test_available_raw_generated_cases_have_deterministic_action_capability(tmp_path):
    profile_path = tmp_path / "raw-wording-profiles.json"
    import_wording_workbook(_raw_workbook_bytes(), profile_path)

    catalog = load_browser_case_catalog(
        progress_path=tmp_path / "missing-progress.json",
        wording_profiles_path=profile_path,
        issuer_mode="sms_otp",
    )

    generated = [
        case
        for case in catalog["cases"]
        if case.get("wording") and case.get("availability", {}).get("enabled")
    ]
    assert generated
    assert all(case["caseImplementation"]["status"] == "completed" for case in generated)
    assert all(case["caseImplementation"]["actionCount"] > 0 for case in generated)
