from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import openpyxl


def load_browser_cases(path: str | Path) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    if "Browser" not in workbook.sheetnames:
        raise ValueError("Workbook does not contain a Browser sheet.")

    sheet = workbook["Browser"]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    cases: list[dict[str, Any]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
        case_id = _text(record.get("ID"))
        if not case_id:
            continue

        function_point = _text(record.get("Function Point"))
        steps = _split_steps(_text(record.get("Steps")))
        expected_text = _text(record.get("Expected Results"))
        cases.append(
            {
                "sheet": "Browser",
                "channel": "browser",
                "id": case_id,
                "system": _text(record.get("System")),
                "module": _text(record.get("Module")),
                "functionPoint": function_point,
                "testPoint": _text(record.get("Test Points")),
                "steps": steps,
                "expected": parse_expected_results(expected_text),
                "automation": _automation_metadata(function_point, _text(record.get("Steps")), expected_text),
                "source": {
                    "actualResult": _text(record.get("Actual Result")),
                    "testDate": _text(record.get("Test Date")),
                    "testers": _text(record.get("Testers")),
                    "remarks": _text(record.get("Remarks")),
                },
            }
        )

    return cases


def export_browser_cases(path: str | Path, output_path: str | Path) -> Path:
    cases = load_browser_cases(path)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(
            {
                "sourceWorkbook": Path(path).name,
                "sheet": "Browser",
                "caseCount": len(cases),
                "cases": cases,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    return output


def parse_expected_results(value: str) -> dict[str, Any]:
    transactions = _parse_transactions(value)
    if transactions:
        return {
            "messages": {},
            "transactions": transactions,
            "prompts": _parse_prompts(value),
            "errors": _parse_errors(value),
        }

    return {
        "messages": _parse_message_expectations(value),
        "transactions": [],
        "prompts": _parse_prompts(value),
        "errors": _parse_errors(value),
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Export Browser SIT cases from the ACS OTP Excel workbook.")
    parser.add_argument("workbook", help="Path to the SIT Excel workbook.")
    parser.add_argument(
        "--output",
        default="sit_cases/pipay_cup_browser_cases.json",
        help="Path for the exported Browser case JSON.",
    )
    args = parser.parse_args(argv)

    output = export_browser_cases(args.workbook, args.output)
    print(f"Exported Browser SIT cases to {output}")
    return 0


def _parse_transactions(value: str) -> list[dict[str, Any]]:
    matches = list(
        re.finditer(
            r"(?im)^\s*(First transaction|Second transaction|Third transaction)\s*:\s*$",
            value,
        )
    )
    if not matches:
        return []

    transactions: list[dict[str, Any]] = []
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(value)
        block = value[start:end]
        transactions.append(
            {
                "label": match.group(1).strip(),
                "messages": _parse_message_expectations(block),
            }
        )
    return transactions


def _parse_message_expectations(value: str) -> dict[str, dict[str, str]]:
    messages: dict[str, dict[str, str]] = {}
    current_message = ""

    for raw_line in value.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        message_match = re.search(r"\b(ARes|CRes|RReq)\s+message\b", line, re.IGNORECASE)
        if message_match:
            current_message = _canonical_message_type(message_match.group(1))
            messages.setdefault(current_message, {})

        if not current_message:
            continue

        reason_match = re.search(r'transStatusReason\s*=\s*"?([0-9A-Za-z]+)"?', line, re.IGNORECASE)
        if reason_match:
            messages[current_message]["transStatusReason"] = reason_match.group(1)
            continue

        status_match = re.search(r'\btransStatus\s*=\s*"?([A-Z])"?', line, re.IGNORECASE)
        if status_match:
            messages[current_message]["transStatus"] = status_match.group(1).upper()
            continue

        eci_match = re.search(r'\bECI\s*=\s*"?([0-9A-Za-z]+)"?', line, re.IGNORECASE)
        if eci_match:
            messages[current_message]["eci"] = eci_match.group(1)
            continue
        if re.search(r"\bECI\s+is\s+null\b", line, re.IGNORECASE):
            messages[current_message]["eci"] = "null"
            continue

        if re.search(r"\bCAVV\s+is\s+not\s+null\b", line, re.IGNORECASE):
            messages[current_message]["cavv"] = "not_null"
            continue
        if re.search(r"\bCAVV\s+is\s+null\b", line, re.IGNORECASE):
            messages[current_message]["cavv"] = "null"

    return messages


def _parse_prompts(value: str) -> list[str]:
    if "Display prompt" not in value and "prompt" not in value.lower():
        return []
    return [_clean_line(line) for line in value.splitlines() if _clean_line(line)]


def _parse_errors(value: str) -> list[dict[str, str]]:
    if not re.search(r"\bReturn Error\b|\bError Code\b", value, re.IGNORECASE):
        return []

    error: dict[str, str] = {}
    for raw_line in value.splitlines():
        line = _clean_line(raw_line)
        code_match = re.search(r"Error Code\s*=\s*([0-9]+)", line, re.IGNORECASE)
        if code_match:
            error["code"] = code_match.group(1)
        description_match = re.search(r"Error Description\s*=\s*(.+)", line, re.IGNORECASE)
        if description_match:
            error["description"] = description_match.group(1).strip()
        detail_match = re.search(r"Error Detail\s*=\s*(.+)", line, re.IGNORECASE)
        if detail_match:
            error["detail"] = detail_match.group(1).strip()
        component_match = re.search(r"Error Component\s*=\s*(.+)", line, re.IGNORECASE)
        if component_match:
            error["component"] = component_match.group(1).strip()
    return [error] if error else []


def _automation_metadata(function_point: str, steps: str, expected: str) -> dict[str, Any]:
    text = f"{function_point}\n{steps}\n{expected}".lower()
    tags: list[str] = []
    tag_patterns = {
        "otp": "otp",
        "challenge": "cres message",
        "rreq": "rreq message",
        "ares": "ares message",
        "invalid_card": "invalid card",
        "retry": "exceeded maximum",
        "pa": "pa",
        "npa": "npa",
        "3ri": "3ri",
        "cancel": "cancel",
        "timeout": "timeout",
        "currency": "currency",
        "error": "return error",
        "prompt": "display prompt",
        "resend": "resend code",
        "expired": "expired",
    }
    for tag, pattern in tag_patterns.items():
        if pattern in text:
            tags.append(tag)

    status = "automatable"
    if any(tag in tags for tag in ["timeout", "resend", "expired"]):
        status = "manual_or_slow"
    elif "prompt" in tags:
        status = "ui_assertion"

    return {
        "status": status,
        "tags": tags,
    }


def _split_steps(value: str) -> list[str]:
    steps: list[str] = []
    for raw_line in value.splitlines():
        line = _clean_line(raw_line)
        if not line:
            continue
        steps.append(line)
    return steps


def _clean_line(value: str) -> str:
    return re.sub(r"^\s*\d+[\.)]\s*", "", value.strip())


def _canonical_message_type(value: str) -> str:
    upper = value.upper()
    if upper == "ARES":
        return "ARes"
    if upper == "CRES":
        return "CRes"
    if upper == "RREQ":
        return "RReq"
    return value


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


if __name__ == "__main__":
    raise SystemExit(main())
