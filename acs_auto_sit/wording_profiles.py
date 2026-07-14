from __future__ import annotations

import json
import re
from hashlib import sha1
from copy import deepcopy
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORDING_PROFILES_PATH = PROJECT_ROOT / "data" / "wording_profiles.json"
DEFAULT_SUPPORTED_LOCALES = ("zh_TW", "en_US", "zh_CN")

WORDING_CASE_SCENARIOS = (
    {
        "baseCaseId": "case23",
        "name": "PA Purchase Authentication",
        "messageCategory": "PA",
        "wordingCode": "SEND_SMS_OTP",
        "scenario": "initial_otp",
    },
    {
        "baseCaseId": "case27",
        "name": "NPA Purchase Authentication",
        "messageCategory": "NPA",
        "wordingCode": "SEND_SMS_OTP",
        "scenario": "initial_otp",
    },
    {
        "baseCaseId": "case31",
        "name": "Incorrect OTP Authentication",
        "messageCategory": "PA",
        "wordingCode": "INCORRECT_SMS_OTP",
        "scenario": "incorrect_otp",
    },
    {
        "baseCaseId": "case35",
        "name": "Resend code Authentication",
        "messageCategory": "PA",
        "wordingCode": "RESEND_SMS_OTP",
        "scenario": "resend_success",
    },
    {
        "baseCaseId": "case39",
        "name": "Resend code before interval",
        "messageCategory": "PA",
        "wordingCode": "RESEND_SMS_GAP_LIMIT",
        "scenario": "resend_gap_limit",
    },
    {
        "baseCaseId": "case43",
        "name": "Resend code limit Authentication",
        "messageCategory": "PA",
        "wordingCode": "RESEND_SMS_LIMIT_EXCEED",
        "scenario": "resend_count_limit",
    },
    {
        "baseCaseId": "case47",
        "name": "Expired OTP Authentication",
        "messageCategory": "PA",
        "wordingCode": "SMS_PASSCODE_EXPIRED",
        "scenario": "expired_otp",
    },
)
EXPECTED_FIELD_ORDER = (
    "challenge_title",
    "challenge_message",
    "challenge_label",
    "second_challenge_label",
    "single_select_option",
    "continue_oob_button",
    "next_button",
    "resend_button",
    "help_label",
    "help_text",
)
REQUIRED_WORDING_FIELDS = ("challenge_title", "challenge_message")

ISSUER_SHEET = "發卡行設定"
WORDING_SHEET = "話術匯入"
ISSUER_REQUIRED_COLUMNS = ("啟用", "發卡行代號", "發卡行名稱", "Issuer Mode", "預設語言")
WORDING_REQUIRED_COLUMNS = (
    "啟用匯入",
    "發卡行代號",
    "Issuer Mode",
    "設備通道",
    "訊息類別",
    "編碼代號",
    "語言代碼",
    "欄位代號",
    "話術內容",
)
RAW_SOURCE_SHEETS = ("SMS", "Email", "OOB", "Single Select")
RAW_REQUIRED_COLUMNS = ("設備通道", "訊息類別", "編碼代號", "語言代碼")
RAW_METADATA_COLUMNS = {*RAW_REQUIRED_COLUMNS, "是否完成"}
RAW_FIELD_KEYS = {
    "驗證訊息標題": "challenge_title",
    "驗證訊息欄位文字": "challenge_message",
    "驗證訊息欄位標籤": "challenge_label",
    "第二組驗證碼標籤": "second_challenge_label",
    "下一步標籤": "next_button",
    "重送驗證標籤": "resend_button",
    "繼續OOB作業標籤": "continue_oob_button",
    "是否需要幫助標籤": "help_label",
    "幫助文字": "help_text",
}
RAW_ISSUER_MODES = (
    "sms_otp",
    "email_otp",
    "direct_oob",
    "selection_sms_oob",
    "selection_sms_email",
    "selection_sms_email_oob",
    "selection_email_oob",
    "default_oob_can_switch_otp",
)


def import_wording_workbook(
    content: bytes,
    destination: Path = DEFAULT_WORDING_PROFILES_PATH,
    *,
    source_file: str = "",
) -> dict[str, Any]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Invalid Excel workbook: {exc}") from exc

    normalized_sheets = {ISSUER_SHEET, WORDING_SHEET}
    available_normalized = normalized_sheets.intersection(workbook.sheetnames)
    source_sheets = [name for name in RAW_SOURCE_SHEETS if name in workbook.sheetnames]
    if normalized_sheets.issubset(workbook.sheetnames):
        issuer_rows = _sheet_records(workbook[ISSUER_SHEET], ISSUER_REQUIRED_COLUMNS)
        wording_rows = _sheet_records(workbook[WORDING_SHEET], WORDING_REQUIRED_COLUMNS)
        issuers = _normalize_issuers(issuer_rows)
        wordings = _normalize_wordings(wording_rows)
        source_format = "normalized"
        source_sheets = list(dict.fromkeys(wording["sourceSheet"] for wording in wordings))
        normalized_row_count = len(wording_rows)
    elif available_normalized:
        missing_sheets = [name for name in (ISSUER_SHEET, WORDING_SHEET) if name not in workbook.sheetnames]
        raise ValueError(f"Missing required sheet(s): {', '.join(missing_sheets)}")
    elif source_sheets:
        issuers, wordings, normalized_row_count = _normalize_raw_workbook(workbook, source_sheets)
        source_format = "challenge_ui_info"
    else:
        expected = ", ".join((ISSUER_SHEET, WORDING_SHEET, *RAW_SOURCE_SHEETS))
        raise ValueError(f"Unsupported wording workbook format. Expected sheets include: {expected}")

    _add_wording_only_issuers(issuers, wordings)
    _apply_supported_locales(issuers, wordings)

    all_locales = {wording["locale"] for wording in wordings}
    normalized = {
        "version": 1,
        "sourceFile": str(source_file or ""),
        "sourceFormat": source_format,
        "sourceSheets": source_sheets,
        "importedAt": datetime.now(timezone.utc).isoformat(),
        "defaultSupportedLocales": list(DEFAULT_SUPPORTED_LOCALES),
        "issuers": issuers,
        "wordings": wordings,
        "summary": {
            "issuerCount": len(issuers),
            "localeCount": len(all_locales),
            "wordingCount": len(wordings),
            "sourceSheetCount": len(source_sheets),
            "normalizedRowCount": normalized_row_count,
        },
    }
    _write_json_atomically(destination, normalized)
    return normalized


def load_wording_profiles(path: Path = DEFAULT_WORDING_PROFILES_PATH) -> dict[str, Any] | None:
    if not path.is_file():
        return None
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("Wording profile JSON must contain an object.")
    return data


def build_localized_wording_cases(
    profiles: dict[str, Any],
    source_cases: list[dict[str, Any]],
    *,
    issuer_id: str = "default",
    issuer_mode: str = "direct_otp",
) -> list[dict[str, Any]]:
    if profiles.get("sourceFormat") == "challenge_ui_info":
        return _build_raw_localized_wording_cases(
            profiles,
            source_cases,
            issuer_id=issuer_id,
            issuer_mode=issuer_mode,
        )

    issuers = profiles.get("issuers") or {}
    issuer = issuers.get(issuer_id) or issuers.get("default") or {}
    selected_issuer_id = str(issuer.get("id") or issuer_id or "default")
    locales = issuer.get("supportedLocales") or profiles.get("defaultSupportedLocales") or DEFAULT_SUPPORTED_LOCALES
    templates = {str(case.get("id") or ""): case for case in source_cases}
    wording_index = _wording_field_index(profiles.get("wordings") or [])
    generated: list[dict[str, Any]] = []

    for scenario in WORDING_CASE_SCENARIOS:
        base_case_id = scenario["baseCaseId"]
        if base_case_id not in templates:
            raise ValueError(f"Missing source wording case template: {base_case_id}")
        for locale in locales:
            locale = str(locale)
            fields = _resolve_wording_fields(
                wording_index,
                selected_issuer_id,
                issuer_mode,
                scenario["messageCategory"],
                scenario["wordingCode"],
                locale,
            )
            missing_fields = [field for field in REQUIRED_WORDING_FIELDS if not fields.get(field)]
            enabled = not missing_fields
            reason = "" if enabled else (
                f"Missing wording fields {', '.join(missing_fields)}: issuer={selected_issuer_id}, "
                f"mode={issuer_mode}, code={scenario['wordingCode']}, locale={locale}."
            )
            expected = deepcopy(templates[base_case_id].get("expected") or {})
            expected["prompts"] = [fields[key] for key in EXPECTED_FIELD_ORDER if key in fields]
            expected["uiFields"] = {key: fields[key] for key in EXPECTED_FIELD_ORDER if key in fields}

            case = deepcopy(templates[base_case_id])
            case.update(
                {
                    "id": f"{base_case_id}_{locale}",
                    "baseCaseId": base_case_id,
                    "functionPoint": f"{scenario['name']} ({locale})",
                    "locale": locale,
                    "browserLanguage": locale.replace("_", "-"),
                    "wordingScenario": scenario["scenario"],
                    "wording": {
                        "issuerId": selected_issuer_id,
                        "issuerMode": issuer_mode,
                        "messageCategory": scenario["messageCategory"],
                        "code": scenario["wordingCode"],
                    },
                    "expected": expected,
                    "availability": {"enabled": enabled, "reason": reason},
                }
            )
            generated.append(case)
    return generated


def _build_raw_localized_wording_cases(
    profiles: dict[str, Any],
    source_cases: list[dict[str, Any]],
    *,
    issuer_id: str,
    issuer_mode: str,
) -> list[dict[str, Any]]:
    from acs_auto_sit.issuer_modes import resolve_issuer_mode

    mode = resolve_issuer_mode(issuer_mode)
    destinations = list(mode.get("destinations") or [])
    templates = {str(case.get("id") or ""): case for case in source_cases}
    groups = _raw_wording_groups(profiles.get("wordings") or [], issuer_id)
    by_sheet = {
        sheet: [group for group in groups if group["sourceSheet"] == sheet]
        for sheet in RAW_SOURCE_SHEETS
    }

    if mode["id"] == "default_oob_can_switch_otp":
        return [
            _raw_switch_case(group, _matching_group(by_sheet["SMS"], group), templates, mode)
            for group in by_sheet["OOB"]
        ]

    if mode["id"].startswith("selection_"):
        selection_groups = by_sheet["Single Select"]
        generated = [
            _raw_case(
                group,
                templates,
                mode,
                flow_kind="selection_page",
                destination="selection",
                stages=[_raw_stage("single_select", group)],
            )
            for group in selection_groups
        ]
        sheet_by_destination = {"sms": "SMS", "email": "Email", "oob": "OOB"}
        for destination in destinations:
            for group in by_sheet[sheet_by_destination[destination]]:
                selection = _matching_group(selection_groups, group)
                generated.append(
                    _raw_case(
                        group,
                        templates,
                        mode,
                        flow_kind="selection_branch",
                        destination=destination,
                        stages=[
                            _raw_stage("single_select", selection),
                            _raw_stage(destination, group),
                        ],
                        extra_missing=[] if selection else [
                            f"Missing Single Select wording for category={group['messageCategory']}, "
                            f"locale={group['locale']}"
                        ],
                    )
                )
        return generated

    sheet_by_destination = {"sms": "SMS", "email": "Email", "oob": "OOB"}
    destination = destinations[0]
    return [
        _raw_case(
            group,
            templates,
            mode,
            flow_kind="direct",
            destination=destination,
            stages=[_raw_stage(destination, group)],
        )
        for group in by_sheet[sheet_by_destination[destination]]
    ]


def _raw_wording_groups(
    wordings: Iterable[dict[str, Any]],
    issuer_id: str,
) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    selected_issuer_id = issuer_id or "default"
    for wording in wordings:
        wording_issuer = str(wording.get("issuerId") or "default")
        if wording_issuer not in {"default", selected_issuer_id}:
            continue
        if str(wording.get("deviceChannel") or "").upper() != "BROWSER":
            continue
        source_sheet = str(wording.get("sourceSheet") or "")
        key = (
            source_sheet,
            str(wording.get("messageCategory") or "").upper(),
            str(wording.get("wordingCode") or "").upper(),
            str(wording.get("locale") or ""),
        )
        group = groups.setdefault(
            key,
            {
                "sourceSheet": key[0],
                "messageCategory": key[1],
                "wordingCode": key[2],
                "locale": key[3],
                "fields": {},
                "completionStatus": str(wording.get("completionStatus") or ""),
            },
        )
        field_key = str(wording.get("fieldKey") or "")
        content = str(wording.get("content") or "")
        if field_key and content:
            group["fields"][field_key] = content
    sheet_order = {sheet: index for index, sheet in enumerate(RAW_SOURCE_SHEETS)}
    return sorted(
        groups.values(),
        key=lambda group: (
            sheet_order.get(group["sourceSheet"], len(sheet_order)),
            group["messageCategory"],
            group["wordingCode"],
            _locale_sort_key(group["locale"]),
        ),
    )


def _raw_case(
    group: dict[str, Any],
    templates: dict[str, dict[str, Any]],
    mode: dict[str, Any],
    *,
    flow_kind: str,
    destination: str,
    stages: list[dict[str, Any]],
    extra_missing: list[str] | None = None,
) -> dict[str, Any]:
    base_case_id = _raw_base_case_id(group)
    template = templates.get(base_case_id) or templates.get("case23") or next(iter(templates.values()), {})
    fields = dict(group["fields"])
    missing = [field for field in REQUIRED_WORDING_FIELDS if not fields.get(field)]
    reasons = list(extra_missing or [])
    if missing:
        reasons.append(
            f"Missing wording fields {', '.join(missing)}: source={group['sourceSheet']}, "
            f"code={group['wordingCode']}, locale={group['locale']}"
        )
    enabled = not reasons
    expected = deepcopy(template.get("expected") or {})
    ordered_keys = [key for key in EXPECTED_FIELD_ORDER if key in fields]
    ordered_keys.extend(sorted(set(fields).difference(ordered_keys)))
    expected["prompts"] = [fields[key] for key in ordered_keys]
    expected["uiFields"] = {key: fields[key] for key in ordered_keys}
    expected["stageUiFields"] = {
        stage["type"]: dict(stage.get("expectedFields") or {})
        for stage in stages
    }

    case = deepcopy(template)
    case.update(
        {
            "id": _raw_case_id(flow_kind, destination, group),
            "baseCaseId": base_case_id,
            "functionPoint": (
                f"{group['sourceSheet']} {group['messageCategory']} "
                f"{group['wordingCode']} ({group['locale']})"
            ),
            "locale": group["locale"],
            "browserLanguage": group["locale"].replace("_", "-"),
            "wordingScenario": _raw_wording_scenario(group["wordingCode"]),
            "wording": {
                "issuerId": "default",
                "issuerMode": mode["id"],
                "sourceSheet": group["sourceSheet"],
                "messageCategory": group["messageCategory"],
                "code": group["wordingCode"],
                "locale": group["locale"],
            },
            "flow": {
                "kind": flow_kind,
                "destination": destination,
                "destinations": list(mode.get("destinations") or []),
                "stages": stages,
            },
            "expected": expected,
            "availability": {"enabled": enabled, "reason": "; ".join(reasons)},
        }
    )
    return case


def _raw_switch_case(
    oob_group: dict[str, Any],
    sms_group: dict[str, Any] | None,
    templates: dict[str, dict[str, Any]],
    mode: dict[str, Any],
) -> dict[str, Any]:
    missing = [] if sms_group else [
        f"Missing SMS wording for category={oob_group['messageCategory']}, locale={oob_group['locale']}"
    ]
    case = _raw_case(
        oob_group,
        templates,
        mode,
        flow_kind="oob_switch_sms",
        destination="sms",
        stages=[_raw_stage("oob", oob_group), _raw_stage("sms", sms_group)],
        extra_missing=missing,
    )
    case["flow"]["switchCreq"] = True
    return case


def _raw_stage(stage_type: str, group: dict[str, Any] | None) -> dict[str, Any]:
    return {
        "type": stage_type,
        "sourceSheet": str(group.get("sourceSheet") or "") if group else "",
        "messageCategory": str(group.get("messageCategory") or "") if group else "",
        "wordingCode": str(group.get("wordingCode") or "") if group else "",
        "locale": str(group.get("locale") or "") if group else "",
        "expectedFields": dict(group.get("fields") or {}) if group else {},
    }


def _matching_group(
    candidates: list[dict[str, Any]],
    target: dict[str, Any],
) -> dict[str, Any] | None:
    for candidate in candidates:
        if (
            candidate["messageCategory"] == target["messageCategory"]
            and candidate["locale"] == target["locale"]
        ):
            return candidate
    return None


def _raw_case_id(flow_kind: str, destination: str, group: dict[str, Any]) -> str:
    if flow_kind == "selection_page":
        prefix = "ui_single_select"
    elif flow_kind == "selection_branch":
        prefix = f"ui_select_{destination}"
    elif flow_kind == "oob_switch_sms":
        prefix = "ui_oob_switch_sms"
    else:
        prefix = f"ui_{destination}"
    parts = (
        prefix,
        group["messageCategory"].lower(),
        group["wordingCode"].lower(),
        group["locale"],
    )
    return "_".join(_id_part(part) for part in parts if part)


def _raw_base_case_id(group: dict[str, Any]) -> str:
    code = group["wordingCode"]
    if "INCORRECT" in code:
        return "case31"
    if "GAP" in code or "INTERVAL" in code:
        return "case39"
    if "LIMIT" in code:
        return "case43"
    if "RESEND" in code:
        return "case35"
    if "EXPIRED" in code:
        return "case47"
    return "case27" if group["messageCategory"] == "NPA" else "case23"


def _raw_wording_scenario(wording_code: str) -> str:
    code = wording_code.upper()
    if "INCORRECT" in code:
        return "incorrect_otp"
    if "GAP" in code or "INTERVAL" in code:
        return "resend_gap_limit"
    if "LIMIT" in code:
        return "resend_count_limit"
    if "RESEND" in code:
        return "resend_success"
    if "EXPIRED" in code:
        return "expired_otp"
    return "initial_challenge"


def _id_part(value: str) -> str:
    return re.sub(r"[^a-zA-Z0-9_]+", "_", value).strip("_")


def _locale_sort_key(locale: str) -> tuple[int, str]:
    try:
        return DEFAULT_SUPPORTED_LOCALES.index(locale), locale
    except ValueError:
        return len(DEFAULT_SUPPORTED_LOCALES), locale


def _wording_field_index(wordings: Iterable[dict[str, Any]]) -> dict[tuple[str, ...], dict[str, str]]:
    index: dict[tuple[str, ...], dict[str, str]] = {}
    for wording in wordings:
        if str(wording.get("deviceChannel") or "").upper() != "BROWSER":
            continue
        key = (
            str(wording.get("issuerId") or "default"),
            str(wording.get("issuerMode") or ""),
            str(wording.get("messageCategory") or "").upper(),
            str(wording.get("wordingCode") or "").upper(),
            str(wording.get("locale") or ""),
        )
        field_key = str(wording.get("fieldKey") or "")
        content = str(wording.get("content") or "")
        if field_key and content:
            index.setdefault(key, {})[field_key] = content
    return index


def _resolve_wording_fields(
    index: dict[tuple[str, ...], dict[str, str]],
    issuer_id: str,
    issuer_mode: str,
    message_category: str,
    wording_code: str,
    locale: str,
) -> dict[str, str]:
    suffix = (issuer_mode, message_category, wording_code, locale)
    fields = dict(index.get(("default", *suffix), {}))
    if issuer_id != "default":
        fields.update(index.get((issuer_id, *suffix), {}))
    return fields


def _sheet_records(sheet: Any, required_columns: Iterable[str]) -> list[dict[str, Any]]:
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        raise ValueError(f"Sheet {sheet.title} is empty.")
    columns = [str(value or "").strip() for value in header]
    missing = [column for column in required_columns if column not in columns]
    if missing:
        raise ValueError(f"Sheet {sheet.title} is missing column(s): {', '.join(missing)}")
    return [dict(zip(columns, row)) for row in rows if any(value not in (None, "") for value in row)]


def _normalize_issuers(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    issuers: dict[str, dict[str, Any]] = {}
    current_id = "default"
    current_name = "預設發卡行"
    for row in rows:
        if not _enabled(row.get("啟用")):
            continue
        current_id = _text(row.get("發卡行代號")) or current_id
        current_name = _text(row.get("發卡行名稱")) or current_name
        issuer = issuers.setdefault(
            current_id,
            {
                "id": current_id,
                "name": current_name,
                "defaultLocale": _text(row.get("預設語言")) or DEFAULT_SUPPORTED_LOCALES[0],
                "issuerModes": [],
                "supportedLocales": [],
            },
        )
        mode = _text(row.get("Issuer Mode"))
        if mode and mode not in issuer["issuerModes"]:
            issuer["issuerModes"].append(mode)
    return issuers


def _normalize_wordings(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    wordings: list[dict[str, Any]] = []
    values_by_key: dict[tuple[str, ...], str] = {}
    for row in rows:
        if not _enabled(row.get("啟用匯入")):
            continue
        normalized = {
            "issuerId": _text(row.get("發卡行代號")) or "default",
            "issuerMode": _required_text(row, "Issuer Mode"),
            "deviceChannel": _required_text(row, "設備通道").upper(),
            "messageCategory": _required_text(row, "訊息類別").upper(),
            "sourceSheet": _required_text(row, "來源頁籤"),
            "wordingCode": _required_text(row, "編碼代號").upper(),
            "locale": _required_text(row, "語言代碼"),
            "fieldKey": _required_text(row, "欄位代號"),
            "content": _required_text(row, "話術內容"),
            "placeholders": list(dict.fromkeys(re.findall(r"\{\d+\}", _text(row.get("話術內容"))))),
            "completionStatus": _text(row.get("完成狀態")),
        }
        option_index = _text(row.get("選項序號"))
        if option_index:
            normalized["optionIndex"] = option_index
        key = tuple(
            normalized[name]
            for name in (
                "issuerId",
                "issuerMode",
                "deviceChannel",
                "messageCategory",
                "sourceSheet",
                "wordingCode",
                "locale",
                "fieldKey",
            )
        ) + (option_index,)
        if key in values_by_key and values_by_key[key] == normalized["content"]:
            continue
        if key in values_by_key:
            raise ValueError(f"Duplicate wording key: {' / '.join(key)}")
        values_by_key[key] = normalized["content"]
        wordings.append(normalized)
    return wordings


def _normalize_raw_workbook(
    workbook: Any,
    source_sheets: list[str],
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]], int]:
    wordings: list[dict[str, Any]] = []
    values_by_key: dict[tuple[str, ...], str] = {}
    row_keys: set[tuple[str, ...]] = set()
    for source_sheet in source_sheets:
        rows = _sheet_records(workbook[source_sheet], RAW_REQUIRED_COLUMNS)
        for row in rows:
            identifiers = [_text(row.get(column)) for column in RAW_REQUIRED_COLUMNS]
            non_metadata_values = [
                value
                for column, value in row.items()
                if column not in RAW_METADATA_COLUMNS and _text(value)
            ]
            if not any(identifiers) and not non_metadata_values:
                continue
            device_channel = _required_text(row, "設備通道").upper()
            if device_channel in {"B", "BRW", "BROWSER"}:
                device_channel = "BROWSER"
            message_category = _required_text(row, "訊息類別").upper()
            wording_code = _required_text(row, "編碼代號").upper()
            locale = _required_text(row, "語言代碼")
            row_key = (source_sheet, device_channel, message_category, wording_code, locale)
            row_keys.add(row_key)
            for source_column, value in row.items():
                content = _text(value)
                if not content or source_column in RAW_METADATA_COLUMNS:
                    continue
                field_key = RAW_FIELD_KEYS.get(source_column) or _raw_field_key(source_column)
                normalized = {
                    "issuerId": "default",
                    "issuerMode": "",
                    "deviceChannel": device_channel,
                    "messageCategory": message_category,
                    "sourceSheet": source_sheet,
                    "wordingCode": wording_code,
                    "locale": locale,
                    "fieldKey": field_key,
                    "sourceColumn": source_column,
                    "content": content,
                    "placeholders": list(dict.fromkeys(re.findall(r"\{\d+\}", content))),
                    "completionStatus": _text(row.get("是否完成")),
                }
                key = (*row_key, field_key)
                previous = values_by_key.get(key)
                if previous == content:
                    continue
                if previous is not None:
                    raise ValueError(f"Duplicate wording key: {' / '.join(key)}")
                values_by_key[key] = content
                wordings.append(normalized)

    issuers = {
        "default": {
            "id": "default",
            "name": "預設發卡行",
            "defaultLocale": DEFAULT_SUPPORTED_LOCALES[0],
            "issuerModes": list(RAW_ISSUER_MODES),
            "supportedLocales": [],
        }
    }
    return issuers, wordings, len(row_keys)


def _raw_field_key(header: str) -> str:
    digest = sha1(header.encode("utf-8")).hexdigest()[:10]
    return f"raw_{digest}"


def _add_wording_only_issuers(
    issuers: dict[str, dict[str, Any]],
    wordings: list[dict[str, Any]],
) -> None:
    for wording in wordings:
        issuer_id = wording["issuerId"]
        issuer = issuers.setdefault(
            issuer_id,
            {
                "id": issuer_id,
                "name": issuer_id,
                "defaultLocale": DEFAULT_SUPPORTED_LOCALES[0],
                "issuerModes": [],
                "supportedLocales": [],
            },
        )
        mode = wording["issuerMode"]
        if mode not in issuer["issuerModes"]:
            issuer["issuerModes"].append(mode)


def _apply_supported_locales(
    issuers: dict[str, dict[str, Any]],
    wordings: list[dict[str, Any]],
) -> None:
    locales_by_issuer: dict[str, set[str]] = {}
    for wording in wordings:
        locales_by_issuer.setdefault(wording["issuerId"], set()).add(wording["locale"])
    for issuer_id, issuer in issuers.items():
        if issuer_id == "default":
            issuer["supportedLocales"] = list(DEFAULT_SUPPORTED_LOCALES)
            continue
        locales = locales_by_issuer.get(issuer_id) or set(DEFAULT_SUPPORTED_LOCALES)
        issuer["supportedLocales"] = _sorted_locales(locales)


def _sorted_locales(locales: Iterable[str]) -> list[str]:
    values = set(locales)
    ordered = [locale for locale in DEFAULT_SUPPORTED_LOCALES if locale in values]
    return ordered + sorted(values.difference(ordered))


def _write_json_atomically(destination: Path, value: dict[str, Any]) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_suffix(f"{destination.suffix}.tmp")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(destination)


def _required_text(row: dict[str, Any], column: str) -> str:
    value = _text(row.get(column))
    if not value:
        raise ValueError(f"Wording row is missing required value: {column}")
    return value


def _enabled(value: Any) -> bool:
    return _text(value).upper() in {"Y", "YES", "TRUE", "1"}


def _text(value: Any) -> str:
    return str(value).strip() if value is not None else ""
